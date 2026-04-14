import openpyxl
import pandas as pd

file_path = r'c:\Users\Mauricio\Documents\GitHub\FadeBooker\Documentación\Documentos\Diccionario de Datos.xlsx'
wb = openpyxl.load_workbook(file_path)

print('Hojas disponibles:')
for sheet_name in wb.sheetnames:
    print(f'  - {sheet_name}')
print('\n' + '='*80 + '\n')

for sheet_name in wb.sheetnames:
    print(f'HOJA: {sheet_name}')
    print('-' * 80)
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print(df.to_string())
    print('\n' + '='*80 + '\n')
